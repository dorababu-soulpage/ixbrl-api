import json
import openpyxl


def convert_excel_to_json(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    json_data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extract data from each row
        data_type, element, *attributes = row

        # Create a dictionary for each row
        json_object = {
            "datatype": data_type,
            "element": element,
            "attributes": [
                attribute for attribute in attributes if attribute is not None
            ],
        }

        # Append the object to the list
        json_data.append(json_object)

    json_data = json.dumps(json_data, indent=4)

    with open("elements.json", "w") as json_file:
        json_file.write(json_data)


# # Convert the Excel sheet to JSON format.
# convert_excel_to_json("data/elements.xlsx")


def format_excel_to_json(excel_file, sheet):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file)

    # Select the desired sheet
    sheet = workbook[sheet]

    # Create a list to store the data
    data = []

    # Get column headers as strings
    headers = [str(cell.value) for cell in sheet[1]]
    # Iterate through rows in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Assuming the first row contains column headers
        # Create a dictionary for each row using column headers as keys
        row_data = dict(zip(headers, row))
        data.append(row_data)

    # Convert the data list to JSON
    json_data = json.dumps(data, indent=4)

    with open("format.json", "w") as json_file:
        json_file.write(json_data)


format_excel_to_json("data/format.xlsx", "Sheet1")
